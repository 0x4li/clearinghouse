from flask import Flask, request, jsonify, send_file, Response
import json, os, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO

app = Flask(__name__)
app.secret_key = "clearinghouse_secret_2024"

# ─── File-based persistence ───────────────────────────────────────────────────
DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
os.makedirs(DATA_DIR, exist_ok=True)
SUBMISSIONS_FILE = os.path.join(DATA_DIR, "submissions.json")
ALLOWED_IDS_FILE = os.path.join(DATA_DIR, "allowed_ids.json")

def load_submissions():
    if os.path.exists(SUBMISSIONS_FILE):
        with open(SUBMISSIONS_FILE, "r") as f:
            return json.load(f)
    return {}

def save_submissions(data):
    with open(SUBMISSIONS_FILE, "w") as f:
        json.dump(data, f)

def load_allowed_ids():
    if os.path.exists(ALLOWED_IDS_FILE):
        with open(ALLOWED_IDS_FILE, "r") as f:
            return set(json.load(f))
    ids = {f"2026{str(i).zfill(4)}" for i in range(1, 183)}
    save_allowed_ids(ids)
    return ids

def save_allowed_ids(ids):
    with open(ALLOWED_IDS_FILE, "w") as f:
        json.dump(list(ids), f)

submissions = load_submissions()
allowed_ids = load_allowed_ids()

TRACKS     = ["SecDev", "NetSec", "Crypto", "GRC"]
CAPACITIES = {"SecDev": 46, "NetSec": 46, "Crypto": 45, "GRC": 45}

# ─── Embedded HTML pages ──────────────────────────────────────────────────────

INDEX_HTML = r"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8"/>
  <meta name="viewport" content="width=device-width, initial-scale=1.0"/>
  <title>Specialty Track Selection</title>
  <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&family=Space+Grotesk:wght@500;700&display=swap');
    *,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
    :root{--navy:#1E3A5F;--blue:#2563EB;--sky:#EFF6FF;--accent:#3B82F6;--green:#16A34A;--red:#DC2626;--gray-1:#F8FAFC;--gray-2:#E2E8F0;--gray-3:#94A3B8;--gray-4:#475569;--text:#0F172A;--radius:12px}
    body{font-family:'Inter',sans-serif;background:var(--gray-1);color:var(--text);min-height:100vh;display:flex;flex-direction:column}
    header{background:var(--navy);color:white;padding:0 2rem;height:64px;display:flex;align-items:center;justify-content:space-between;position:sticky;top:0;z-index:100;box-shadow:0 2px 12px rgba(0,0,0,.08)}
    header .brand{font-family:'Space Grotesk',sans-serif;font-size:1.1rem;font-weight:700;display:flex;align-items:center;gap:10px}
    header .brand .dot{width:8px;height:8px;border-radius:50%;background:#60A5FA}
    header nav a{color:#93C5FD;text-decoration:none;font-size:.85rem;font-weight:500;padding:6px 12px;border-radius:6px;transition:background .2s}
    header nav a:hover{background:rgba(255,255,255,.1);color:white}
    .hero{background:linear-gradient(135deg,var(--navy) 0%,#1e4080 60%,#2563EB 100%);color:white;padding:3.5rem 2rem 3rem;text-align:center}
    .hero h1{font-family:'Space Grotesk',sans-serif;font-size:clamp(1.6rem,4vw,2.4rem);font-weight:700;margin-bottom:.6rem}
    .hero p{color:#BFDBFE;font-size:.95rem;max-width:480px;margin:0 auto;line-height:1.6}
    main{max-width:720px;margin:2.5rem auto;padding:0 1.5rem 4rem;width:100%}
    .card{background:white;border-radius:var(--radius);box-shadow:0 1px 3px rgba(0,0,0,.06),0 4px 16px rgba(0,0,0,.04);padding:2rem;margin-bottom:1.5rem}
    .card-title{font-family:'Space Grotesk',sans-serif;font-size:1rem;font-weight:700;color:var(--navy);margin-bottom:1.25rem;padding-bottom:.75rem;border-bottom:2px solid var(--sky);display:flex;align-items:center;gap:10px}
    .card-title .icon{width:28px;height:28px;background:var(--sky);border-radius:7px;display:flex;align-items:center;justify-content:center;font-size:14px}
    .form-row{display:grid;grid-template-columns:1fr 1fr;gap:1rem}
    .form-group{display:flex;flex-direction:column;gap:6px}
    .form-group label{font-size:.8rem;font-weight:600;color:var(--gray-4);text-transform:uppercase;letter-spacing:.5px}
    .form-group input{border:1.5px solid var(--gray-2);border-radius:8px;padding:10px 14px;font-size:.95rem;font-family:'Inter',sans-serif;color:var(--text);transition:border-color .2s,box-shadow .2s}
    .form-group input:focus{border-color:var(--accent);box-shadow:0 0 0 3px rgba(59,130,246,.12)}
    .scores-grid{display:grid;grid-template-columns:repeat(4,1fr);gap:.75rem}
    .score-box{background:var(--sky);border:2px solid transparent;border-radius:10px;padding:.9rem .5rem;text-align:center;transition:border-color .2s}
    .score-box:focus-within{border-color:var(--accent)}
    .score-box .sem-label{font-size:.7rem;font-weight:700;color:var(--navy);text-transform:uppercase;letter-spacing:.8px;margin-bottom:6px}
    .score-box input{border:none!important;background:transparent!important;text-align:center;font-size:1.3rem;font-weight:600;color:var(--navy);width:100%;box-shadow:none!important;padding:0!important}
    .score-box .hint{font-size:.65rem;color:var(--gray-3);margin-top:4px}
    .rank-instructions{font-size:.82rem;color:var(--gray-3);margin-bottom:1rem}
    .track-list{display:flex;flex-direction:column;gap:10px}
    .track-item{display:flex;align-items:center;gap:14px;background:var(--gray-1);border:1.5px solid var(--gray-2);border-radius:10px;padding:12px 16px;cursor:grab;transition:all .18s;user-select:none}
    .track-item:active{cursor:grabbing}
    .track-item.dragging{opacity:.45;border-style:dashed}
    .track-item.drag-over{border-color:var(--accent);background:var(--sky);transform:scale(1.01)}
    .rank-badge{width:28px;height:28px;background:var(--navy);color:white;border-radius:50%;font-size:.8rem;font-weight:700;display:flex;align-items:center;justify-content:center;flex-shrink:0}
    .track-icon{width:36px;height:36px;border-radius:9px;display:flex;align-items:center;justify-content:center;font-size:18px;flex-shrink:0}
    .track-item[data-track="SecDev"] .track-icon{background:#DCFCE7}
    .track-item[data-track="NetSec"] .track-icon{background:#DBEAFE}
    .track-item[data-track="Crypto"] .track-icon{background:#FCE7F3}
    .track-item[data-track="GRC"]    .track-icon{background:#FEF3C7}
    .track-info .track-name{font-weight:600;font-size:.9rem;color:var(--text)}
    .track-info .track-desc{font-size:.75rem;color:var(--gray-3)}
    .drag-handle{margin-left:auto;color:var(--gray-3);font-size:18px}
    .submit-btn{width:100%;background:var(--blue);color:white;border:none;border-radius:10px;padding:15px;font-size:1rem;font-weight:600;font-family:'Inter',sans-serif;cursor:pointer;transition:all .2s}
    .submit-btn:hover{background:#1D4ED8;transform:translateY(-1px);box-shadow:0 4px 12px rgba(37,99,235,.3)}
    .submit-btn:disabled{background:var(--gray-3);cursor:not-allowed;transform:none;box-shadow:none}
    #toast{position:fixed;bottom:2rem;left:50%;transform:translateX(-50%) translateY(80px);background:white;border-radius:12px;padding:14px 22px;box-shadow:0 8px 32px rgba(0,0,0,.12);display:flex;align-items:center;gap:8px;z-index:200;transition:transform .3s}
    #toast.show{transform:translateX(-50%) translateY(0)}
    #toast.success{border-left:4px solid var(--green)}
    #toast.error{border-left:4px solid var(--red)}
    .spinner{width:18px;height:18px;border:2px solid rgba(255,255,255,.4);border-top-color:white;border-radius:50%;animation:spin .6s linear infinite}
    @keyframes spin{to{transform:rotate(360deg)}}
    #success-screen{display:none;text-align:center;padding:3rem 2rem}
    #success-screen .check{font-size:4rem;margin-bottom:1rem}
    #success-screen h2{font-family:'Space Grotesk',sans-serif;font-size:1.6rem;color:var(--navy);margin-bottom:.5rem}
    #success-screen p{color:var(--gray-3)}
    @media(max-width:520px){.form-row{grid-template-columns:1fr}.scores-grid{grid-template-columns:repeat(2,1fr)}}
  </style>
</head>
<body>
<header>
  <div class="brand"><div class="dot"></div>University Clearinghouse</div>
</header>
<div class="hero">
  <h1>Specialty Track Selection</h1>
  <p>Submit your academic record and rank your preferred specializations. Assignments are determined by merit and availability.</p>
</div>
<main>
  <div id="success-screen" class="card">
    <div class="check">✅</div>
    <h2>Preferences Recorded</h2>
    <p>Your submission has been received. Final track assignments will be published after the deadline by the administrator.</p>
  </div>
  <div id="form-wrapper">
    <div class="card">
      <div class="card-title"><div class="icon">🪪</div>Student Information</div>
      <div class="form-row" style="margin-bottom:1rem">
        <div class="form-group"><label>Student ID</label><input type="text" id="student_id" placeholder="e.g. 20260001" maxlength="20"/></div>
        <div class="form-group"><label>Full Name</label><input type="text" id="name" placeholder="Your full name"/></div>
      </div>
    </div>
    <div class="card">
      <div class="card-title"><div class="icon">📊</div>Academic Record (0 – 20)</div>
      <div class="scores-grid">
        <div class="score-box"><div class="sem-label">Semester 1</div><input type="number" id="s1" min="0" max="20" step="0.01" placeholder="—"/><div class="hint">/ 20</div></div>
        <div class="score-box"><div class="sem-label">Semester 2</div><input type="number" id="s2" min="0" max="20" step="0.01" placeholder="—"/><div class="hint">/ 20</div></div>
        <div class="score-box"><div class="sem-label">Semester 3</div><input type="number" id="s3" min="0" max="20" step="0.01" placeholder="—"/><div class="hint">/ 20</div></div>
        <div class="score-box"><div class="sem-label">Semester 4</div><input type="number" id="s4" min="0" max="20" step="0.01" placeholder="—"/><div class="hint">/ 20</div></div>
      </div>
    </div>
    <div class="card">
      <div class="card-title"><div class="icon">🏆</div>Rank Your Specializations</div>
      <div class="rank-instructions">☰ Drag to reorder — top = most preferred</div>
      <div class="track-list" id="track-list">
        <div class="track-item" data-track="SecDev" draggable="true"><div class="rank-badge">1</div><div class="track-icon">🔐</div><div class="track-info"><div class="track-name">SecDev</div><div class="track-desc">Secure Development</div></div></div>
        <div class="track-item" data-track="NetSec" draggable="true"><div class="rank-badge">2</div><div class="track-icon">🌐</div><div class="track-info"><div class="track-name">NetSec</div><div class="track-desc">Network Security</div></div></div>
        <div class="track-item" data-track="Crypto" draggable="true"><div class="rank-badge">3</div><div class="track-icon">🔑</div><div class="track-info"><div class="track-name">Crypto</div><div class="track-desc">Cryptography</div></div></div>
        <div class="track-item" data-track="GRC" draggable="true"><div class="rank-badge">4</div><div class="track-icon">📋</div><div class="track-info"><div class="track-name">GRC</div><div class="track-desc">Governance & Compliance</div></div></div>
      </div>
    </div>
    <button class="submit-btn" id="submit-btn" onclick="submitForm()"><span id="btn-label">Submit Preferences</span></button>
  </div>
</main>
<div id="toast"><span class="toast-icon" id="toast-icon"></span><span id="toast-msg"></span></div>
<script>
  const list = document.getElementById('track-list');
  let dragged = null;
  list.addEventListener('dragstart', e => { dragged = e.target.closest('.track-item'); setTimeout(() => dragged.classList.add('dragging'), 0); });
  list.addEventListener('dragend', e => { e.target.closest('.track-item').classList.remove('dragging'); updateRanks(); });
  list.addEventListener('dragover', e => {
    e.preventDefault();
    const target = e.target.closest('.track-item');
    if (target && target !== dragged) {
      document.querySelectorAll('.track-item').forEach(i => i.classList.remove('drag-over'));
      target.classList.add('drag-over');
      const rect = target.getBoundingClientRect();
      list.insertBefore(dragged, (e.clientY - rect.top) > rect.height / 2 ? target.nextSibling : target);
    }
  });
  list.addEventListener('dragleave', e => e.target.closest?.('.track-item')?.classList.remove('drag-over'));
  list.addEventListener('drop', e => { e.preventDefault(); document.querySelectorAll('.track-item').forEach(i => i.classList.remove('drag-over')); });
  function updateRanks() { document.querySelectorAll('.track-item').forEach((item, i) => { item.querySelector('.rank-badge').textContent = i + 1; }); }
  async function submitForm() {
    const btn = document.getElementById('submit-btn');
    const label = document.getElementById('btn-label');
    const student_id = document.getElementById('student_id').value.trim();
    const name = document.getElementById('name').value.trim();
    const s1 = parseFloat(document.getElementById('s1').value);
    const s2 = parseFloat(document.getElementById('s2').value);
    const s3 = parseFloat(document.getElementById('s3').value);
    const s4 = parseFloat(document.getElementById('s4').value);
    const choices = [...document.querySelectorAll('.track-item')].map(i => i.dataset.track);
    if (!student_id || !name) { showToast('error','⚠️','Please enter your Student ID and name.'); return; }
    if ([s1,s2,s3,s4].some(isNaN)) { showToast('error','⚠️','Please fill in all four semester scores.'); return; }
    if ([s1,s2,s3,s4].some(v => v < 0 || v > 20)) { showToast('error','⚠️','Scores must be between 0 and 20.'); return; }
    btn.disabled = true;
    label.innerHTML = '<div class="spinner"></div> Submitting…';
    try {
      const res = await fetch('/api/submit', { method:'POST', headers:{'Content-Type':'application/json'}, body: JSON.stringify({student_id,name,s1,s2,s3,s4,choices}) });
      const data = await res.json();
      if (data.success) { document.getElementById('form-wrapper').style.display='none'; document.getElementById('success-screen').style.display='block'; }
      else { showToast('error','✗',data.error); btn.disabled=false; label.textContent='Submit Preferences'; }
    } catch { showToast('error','✗','Network error. Please try again.'); btn.disabled=false; label.textContent='Submit Preferences'; }
  }
  let toastTimer;
  function showToast(type, icon, msg) {
    const t = document.getElementById('toast');
    document.getElementById('toast-icon').textContent = icon;
    document.getElementById('toast-msg').textContent = msg;
    t.className = `show ${type}`;
    clearTimeout(toastTimer);
    toastTimer = setTimeout(() => t.className = type, 4000);
  }
</script>
</body></html>"""

ADMIN_HTML = r"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8"/>
  <meta name="viewport" content="width=device-width, initial-scale=1.0"/>
  <title>Admin Panel — University Clearinghouse</title>
  <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&family=Space+Grotesk:wght@500;700&display=swap');
    *,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
    :root{--navy:#1E3A5F;--blue:#2563EB;--sky:#EFF6FF;--green:#16A34A;--red:#DC2626;--gold:#D97706;--gray-1:#F8FAFC;--gray-2:#E2E8F0;--gray-3:#94A3B8;--gray-4:#475569;--text:#0F172A}
    body{font-family:'Inter',sans-serif;background:var(--gray-1);color:var(--text);min-height:100vh}
    header{background:var(--navy);color:white;padding:0 2rem;height:64px;display:flex;align-items:center;justify-content:space-between;position:sticky;top:0;z-index:100;box-shadow:0 2px 12px rgba(0,0,0,.08)}
    header .brand{font-family:'Space Grotesk',sans-serif;font-size:1.1rem;font-weight:700;display:flex;align-items:center;gap:10px}
    header .brand .badge{background:#F59E0B;color:#1C1917;font-size:.65rem;font-weight:700;padding:2px 8px;border-radius:99px;text-transform:uppercase;letter-spacing:.5px}
    header nav a{color:#93C5FD;text-decoration:none;font-size:.85rem;font-weight:500;padding:6px 12px;border-radius:6px;transition:background .2s}
    header nav a:hover{background:rgba(255,255,255,.1);color:white}
    main{max-width:1100px;margin:2.5rem auto;padding:0 1.5rem 4rem}
    h1{font-family:'Space Grotesk',sans-serif;font-size:1.6rem;font-weight:700;color:var(--navy);margin-bottom:.4rem}
    .subtitle{color:var(--gray-3);font-size:.9rem;margin-bottom:2rem}
    .stats-bar{display:grid;grid-template-columns:repeat(auto-fit,minmax(160px,1fr));gap:1rem;margin-bottom:2rem}
    .stat-card{background:white;border-radius:12px;padding:1.2rem 1.4rem;box-shadow:0 1px 3px rgba(0,0,0,.06);border-left:4px solid var(--blue)}
    .stat-card.green{border-color:var(--green)}.stat-card.gold{border-color:var(--gold)}
    .stat-card .stat-num{font-family:'Space Grotesk',sans-serif;font-size:2rem;font-weight:700;color:var(--navy);line-height:1}
    .stat-card .stat-label{font-size:.75rem;color:var(--gray-3);margin-top:4px}
    .card{background:white;border-radius:12px;box-shadow:0 1px 3px rgba(0,0,0,.06);padding:1.5rem;margin-bottom:1.5rem}
    .card-title{font-family:'Space Grotesk',sans-serif;font-size:.95rem;font-weight:700;color:var(--navy);margin-bottom:1rem;padding-bottom:.75rem;border-bottom:2px solid var(--sky)}
    .upload-zone{border:2px dashed var(--gray-2);border-radius:10px;padding:2rem;text-align:center;cursor:pointer;transition:all .2s}
    .upload-zone:hover{border-color:var(--blue);background:var(--sky)}
    .upload-zone input{display:none}
    .upload-zone .icon{font-size:2.5rem;margin-bottom:.5rem}
    .upload-zone p{color:var(--gray-3);font-size:.85rem}
    .upload-zone strong{color:var(--navy)}
    .btn{border:none;border-radius:8px;padding:10px 20px;font-size:.9rem;font-weight:600;font-family:'Inter',sans-serif;cursor:pointer;transition:all .18s;display:inline-flex;align-items:center;gap:6px}
    .btn-primary{background:var(--blue);color:white}.btn-primary:hover{background:#1D4ED8;transform:translateY(-1px)}
    .btn-green{background:#DCFCE7;color:var(--green)}.btn-green:hover{background:#BBF7D0}
    .btn:disabled{opacity:.4;cursor:not-allowed;transform:none}
    .action-row{display:flex;gap:.75rem;flex-wrap:wrap;margin-top:1rem}
    .method-tabs{display:flex;gap:8px;margin-bottom:1rem}
    .method-tab{padding:8px 16px;border-radius:8px;border:2px solid var(--gray-2);font-size:.85rem;font-weight:600;cursor:pointer;transition:all .18s;background:white;color:var(--gray-4)}
    .method-tab.active{border-color:var(--blue);background:var(--sky);color:var(--blue)}
    .quota-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:1rem}
    .quota-card{background:var(--gray-1);border-radius:10px;padding:1rem}
    .quota-card .track-name{font-weight:600;font-size:.9rem;margin-bottom:6px}
    .quota-bar{height:8px;background:var(--gray-2);border-radius:99px;overflow:hidden;margin-bottom:4px}
    .quota-bar .fill{height:100%;border-radius:99px;transition:width .6s}
    .secdev-fill{background:#22C55E}.netsec-fill{background:#3B82F6}.crypto-fill{background:#EC4899}.grc-fill{background:#F59E0B}
    .quota-nums{font-size:.75rem;color:var(--gray-3)}
    .table-wrap{overflow-x:auto;margin-top:1.5rem;border-radius:10px;border:1px solid var(--gray-2)}
    table{width:100%;border-collapse:collapse;font-size:.85rem}
    thead th{background:var(--navy);color:white;padding:11px 14px;text-align:left;font-weight:600;white-space:nowrap}
    tbody tr:nth-child(even){background:var(--sky)}
    tbody tr:hover{background:#DBEAFE}
    tbody td{padding:9px 14px;border-bottom:1px solid var(--gray-2);white-space:nowrap}
    .badge{display:inline-block;padding:3px 10px;border-radius:99px;font-size:.72rem;font-weight:600}
    .badge-secdev{background:#DCFCE7;color:#15803D}.badge-netsec{background:#DBEAFE;color:#1D4ED8}.badge-crypto{background:#FCE7F3;color:#BE185D}.badge-grc{background:#FEF3C7;color:#B45309}
    .badge-yes{background:#DCFCE7;color:#15803D}.badge-no{background:#FEE2E2;color:#B91C1C}
    #upload-status{font-size:.85rem;margin-top:.5rem;color:var(--green)}
    .empty-state{text-align:center;padding:3rem;color:var(--gray-3)}
  </style>
</head>
<body>
<header>
  <div class="brand">University Clearinghouse <span class="badge">Admin</span></div>
  <nav><a href="/">← Student Form</a></nav>
</header>
<main>
  <h1>Allocation Control Panel</h1>
  <p class="subtitle">Monitor submissions, run the allocation engine, and download final rosters.</p>
  <div class="stats-bar">
    <div class="stat-card"><div class="stat-num" id="stat-submitted">—</div><div class="stat-label">Submitted</div></div>
    <div class="stat-card gold"><div class="stat-num" id="stat-remaining">—</div><div class="stat-label">Remaining</div></div>
    <div class="stat-card green"><div class="stat-num" id="stat-total">—</div><div class="stat-label">Total Students</div></div>
  </div>
  <div class="card">
    <div class="card-title">📁 Upload Master Student List</div>
    <div class="upload-zone" onclick="document.getElementById('file-input').click()">
      <div class="icon">📊</div>
      <input type="file" id="file-input" accept=".xlsx,.xls,.csv" onchange="uploadFile(this)"/>
      <p><strong>Click to upload</strong> your .xlsx or .csv file</p>
      <p>First column must contain Student IDs (with header row)</p>
    </div>
    <div id="upload-status"></div>
  </div>
  <div class="card">
    <div class="card-title">⚙️ Run Allocation Engine</div>
    <div class="method-tabs">
      <div class="method-tab active" id="tab-1" onclick="selectMethod(1)">Method 1 — Late Performance (S3, S4)</div>
      <div class="method-tab" id="tab-2" onclick="selectMethod(2)">Method 2 — Overall GPA (S1–S4)</div>
    </div>
    <div class="action-row">
      <button class="btn btn-primary" onclick="runAllocation()">▶ Run Allocation</button>
      <button class="btn btn-green" id="btn-export" onclick="exportRoster()" disabled>⬇ Download Excel Roster</button>
    </div>
    <div class="quota-grid" id="quota-grid" style="margin-top:1.5rem;display:none">
      <div class="quota-card"><div class="track-name">🔐 SecDev</div><div class="quota-bar"><div class="fill secdev-fill" id="bar-secdev" style="width:0%"></div></div><div class="quota-nums" id="num-secdev">0 / 46</div></div>
      <div class="quota-card"><div class="track-name">🌐 NetSec</div><div class="quota-bar"><div class="fill netsec-fill" id="bar-netsec" style="width:0%"></div></div><div class="quota-nums" id="num-netsec">0 / 46</div></div>
      <div class="quota-card"><div class="track-name">🔑 Crypto</div><div class="quota-bar"><div class="fill crypto-fill" id="bar-crypto" style="width:0%"></div></div><div class="quota-nums" id="num-crypto">0 / 45</div></div>
      <div class="quota-card"><div class="track-name">📋 GRC</div><div class="quota-bar"><div class="fill grc-fill" id="bar-grc" style="width:0%"></div></div><div class="quota-nums" id="num-grc">0 / 45</div></div>
    </div>
  </div>
  <div class="card">
    <div class="card-title">📋 Allocation Results</div>
    <div id="results-area"><div class="empty-state">Run the allocation engine above to see results here.</div></div>
  </div>
</main>
<script>
  let currentMethod = 1;
  async function loadStats() {
    try {
      const r = await fetch('/api/stats');
      const d = await r.json();
      document.getElementById('stat-submitted').textContent = d.total_submitted;
      document.getElementById('stat-remaining').textContent = d.remaining;
      document.getElementById('stat-total').textContent = d.total_allowed;
    } catch {}
  }
  loadStats();
  setInterval(loadStats, 8000);
  function selectMethod(m) {
    currentMethod = m;
    document.getElementById('tab-1').classList.toggle('active', m===1);
    document.getElementById('tab-2').classList.toggle('active', m===2);
    document.getElementById('btn-export').disabled = true;
    document.getElementById('quota-grid').style.display = 'none';
    document.getElementById('results-area').innerHTML = '<div class="empty-state">Run allocation to see results.</div>';
  }
  async function uploadFile(input) {
    const file = input.files[0]; if (!file) return;
    const fd = new FormData(); fd.append('file', file);
    const status = document.getElementById('upload-status');
    status.textContent = 'Uploading…'; status.style.color = 'var(--green)';
    try {
      const r = await fetch('/api/upload_ids', {method:'POST', body:fd});
      const d = await r.json();
      if (d.success) { status.textContent = `✓ Loaded ${d.loaded} student IDs from ${file.name}`; loadStats(); }
      else { status.style.color='var(--red)'; status.textContent = d.error; }
    } catch { status.style.color='var(--red)'; status.textContent='Upload failed.'; }
  }
  async function runAllocation() {
    const res = await fetch(`/api/allocate/${currentMethod}`);
    const data = await res.json();
    const caps = {SecDev:46, NetSec:46, Crypto:45, GRC:45};
    for (const [track, cap] of Object.entries(caps)) {
      const key = track.toLowerCase().replace(' ','');
      const filled = data.track_counts[track] || 0;
      document.getElementById(`bar-${key}`).style.width = `${(filled/cap)*100}%`;
      document.getElementById(`num-${key}`).textContent = `${filled} / ${cap}`;
    }
    document.getElementById('quota-grid').style.display = 'grid';
    const trackBadge = t => { const cls={SecDev:'secdev',NetSec:'netsec',Crypto:'crypto',GRC:'grc'}[t]||''; return `<span class="badge badge-${cls}">${t}</span>`; };
    const rows = data.assignments.map(a => `<tr><td>${a.rank}</td><td><strong>${a.student_id}</strong></td><td>${a.name||'—'}</td><td>${a.s1}</td><td>${a.s2}</td><td>${a.s3}</td><td>${a.s4}</td><td>${a.metric}</td><td>${a.first_choice}</td><td>${trackBadge(a.assigned_track)}</td><td><span class="badge ${a.got_first_choice?'badge-yes':'badge-no'}">${a.got_first_choice?'Yes':'No'}</span></td></tr>`).join('');
    document.getElementById('results-area').innerHTML = `<p style="color:var(--gray-3);font-size:.82rem;margin-bottom:.75rem">Method ${data.method} · Metric: ${data.metric_label} · ${data.assignments.length} total assignments</p><div class="table-wrap"><table><thead><tr><th>Rank</th><th>Student ID</th><th>Name</th><th>S1</th><th>S2</th><th>S3</th><th>S4</th><th>Metric</th><th>1st Choice</th><th>Assigned</th><th>Got 1st?</th></tr></thead><tbody>${rows}</tbody></table></div>`;
    document.getElementById('btn-export').disabled = false;
    loadStats();
  }
  function exportRoster() { window.location.href = `/api/export/${currentMethod}`; }
</script>
</body></html>"""

# ─── Routes ────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return Response(INDEX_HTML, mimetype="text/html")

@app.route("/admin")
def admin():
    return Response(ADMIN_HTML, mimetype="text/html")

@app.route("/api/submit", methods=["POST"])
def submit():
    data = request.get_json()
    student_id = str(data.get("student_id", "")).strip()
    name       = str(data.get("name", "")).strip()

    if student_id not in allowed_ids:
        return jsonify({"success": False, "error": "Invalid Student ID. You are not registered in the system."}), 403
    if student_id in submissions:
        return jsonify({"success": False, "error": "You have already submitted your preferences."}), 409

    try:
        s1, s2, s3, s4 = float(data["s1"]), float(data["s2"]), float(data["s3"]), float(data["s4"])
    except (KeyError, ValueError):
        return jsonify({"success": False, "error": "Invalid semester scores."}), 400

    for s in [s1, s2, s3, s4]:
        if not (0 <= s <= 20):
            return jsonify({"success": False, "error": "Scores must be between 0 and 20."}), 400

    choices = data.get("choices", [])
    if sorted(choices) != sorted(TRACKS):
        return jsonify({"success": False, "error": "Invalid track choices. Rank all 4 tracks."}), 400

    submissions[student_id] = {
        "student_id": student_id, "name": name,
        "s1": s1, "s2": s2, "s3": s3, "s4": s4,
        "choices": choices,
        "metric1": (s3 + s4) / 2,
        "metric2": (s1 + s2 + s3 + s4) / 4,
    }
    save_submissions(submissions)
    return jsonify({"success": True, "message": "Preferences submitted successfully!"})


@app.route("/api/stats")
def stats():
    return jsonify({
        "total_submitted": len(submissions),
        "total_allowed": len(allowed_ids),
        "remaining": len(allowed_ids) - len(submissions),
    })


def run_allocation_logic(method):
    metric_key = "metric1" if method == 1 else "metric2"
    sorted_students = sorted(submissions.values(), key=lambda x: x[metric_key], reverse=True)
    counts = {t: 0 for t in TRACKS}
    assignments = []
    for rank, student in enumerate(sorted_students, start=1):
        assigned_track = None
        for choice in student["choices"]:
            if counts[choice] < CAPACITIES[choice]:
                assigned_track = choice
                counts[choice] += 1
                break
        assignments.append({
            "rank": rank,
            "student_id": student["student_id"],
            "name": student["name"],
            "s1": student["s1"], "s2": student["s2"],
            "s3": student["s3"], "s4": student["s4"],
            "metric": round(student[metric_key], 4),
            "first_choice": student["choices"][0],
            "assigned_track": assigned_track or "Unassigned",
            "got_first_choice": assigned_track == student["choices"][0],
        })
    return assignments, counts


@app.route("/api/allocate/<int:method>")
def allocate(method):
    if method not in (1, 2):
        return jsonify({"error": "Method must be 1 or 2"}), 400
    assignments, counts = run_allocation_logic(method)
    return jsonify({
        "method": method,
        "metric_label": "Avg(S3,S4)" if method == 1 else "Avg(S1-S4)",
        "assignments": assignments,
        "track_counts": counts,
    })


@app.route("/api/export/<int:method>")
def export(method):
    if method not in (1, 2):
        return jsonify({"error": "Method must be 1 or 2"}), 400

    assignments, _ = run_allocation_logic(method)
    label = "S3-S4 Performance" if method == 1 else "Overall GPA"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Roster Method {method}"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    accent_fill = PatternFill("solid", fgColor="E8F0FE")
    green_fill  = PatternFill("solid", fgColor="D4EDDA")
    orange_fill = PatternFill("solid", fgColor="FFF3CD")
    center      = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"), bottom=Side(style="thin"))
    track_colors = {"SecDev":"C8E6C9","NetSec":"BBDEFB","Crypto":"F8BBD0","GRC":"FFE0B2","Unassigned":"F5F5F5"}

    ws.merge_cells("A1:K1")
    ws["A1"] = f"University Clearinghouse — Allocation Roster | Method {method}: {label}"
    ws["A1"].font = Font(bold=True, size=14, color="1E3A5F")
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 28

    headers = ["Rank","Student ID","Name","S1","S2","S3","S4",f"Metric ({label})","1st Choice","Assigned Track","Got 1st Choice?"]
    col_widths = [8,14,22,7,7,7,7,18,14,16,16]

    for col,(h,w) in enumerate(zip(headers,col_widths),start=1):
        cell = ws.cell(row=2,column=col,value=h)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = center; cell.border = thin_border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    for row_idx, a in enumerate(assignments, start=3):
        row_fill = PatternFill("solid", fgColor=track_colors.get(a["assigned_track"],"FFFFFF"))
        values = [a["rank"],a["student_id"],a["name"],a["s1"],a["s2"],a["s3"],a["s4"],
                  a["metric"],a["first_choice"],a["assigned_track"],"Yes" if a["got_first_choice"] else "No"]
        for col,val in enumerate(values,start=1):
            cell = ws.cell(row=row_idx,column=col,value=val)
            cell.alignment = center; cell.border = thin_border
            if col==10: cell.fill = row_fill
            elif col==11: cell.fill = green_fill if a["got_first_choice"] else orange_fill
            else: cell.fill = accent_fill if row_idx%2==0 else PatternFill()

    ws2 = wb.create_sheet("Summary")
    for i,h in enumerate(["Track","Capacity","Assigned"],start=1):
        ws2.cell(1,i,h).font = Font(bold=True)
    for i,track in enumerate(TRACKS,start=2):
        cnt = sum(1 for a in assignments if a["assigned_track"]==track)
        ws2.cell(i,1,track); ws2.cell(i,2,CAPACITIES[track]); ws2.cell(i,3,cnt)

    buf = BytesIO()
    wb.save(buf); buf.seek(0)
    filename = f"Roster_Method{method}_{label.replace(' ','_')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/api/upload_ids", methods=["POST"])
def upload_ids():
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "No file provided"}), 400
    new_ids = set()
    fname = file.filename.lower()
    if fname.endswith(".xlsx") or fname.endswith(".xls"):
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]: new_ids.add(str(row[0]).strip())
    elif fname.endswith(".csv"):
        import csv, io
        content = file.read().decode("utf-8")
        reader = csv.reader(io.StringIO(content))
        next(reader, None)
        for row in reader:
            if row: new_ids.add(str(row[0]).strip())
    else:
        return jsonify({"error": "Only .xlsx or .csv supported"}), 400
    allowed_ids.clear(); allowed_ids.update(new_ids)
    save_allowed_ids(allowed_ids)
    return jsonify({"success": True, "loaded": len(allowed_ids)})


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
